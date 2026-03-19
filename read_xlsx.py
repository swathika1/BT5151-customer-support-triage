from openpyxl import load_workbook

try:
    wb = load_workbook('task_distribution.xlsx')
    ws = wb.active
    
    print("Task Distribution:")
    print("="*80)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        print(" | ".join(str(cell) if cell is not None else "" for cell in row))
except Exception as e:
    print(f"Error: {e}")
